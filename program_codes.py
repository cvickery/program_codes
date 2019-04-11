#! /usr/local/bin/python3
""" Create a spreadsheet giving information for each NYSDOE-approved academic program
    at Queens College.

  Scrapes the QC information from the NYS DOE website.

  The spreadsheet lines consist of the following columns. The combination of values listed as
  the primary key are guaranteed to be unique.

  Primary Key
    program_code, program_name, HEGIS_code, award, institution
  Fields
    certificate_licenses_titles_types, TAP_eligible, APTS_eligible, WTA_eligible, accreditation,
    Unit Code

April 2019:
      Unit Code is new. “Applications for program revisions, title changes and program
      discontinuances should be submitted to the NYSED office that originally registered the
      program.”
        OP:   Office of the Professions
        OCUE: Office of College and University Evaluation
      The Unit Code is on the same page as the program code, so the functionality of the bash
      script that used to generate a program_codes.out file has now been folded into this script in
      order to pick up this new information.

      That is, this is a two-step process:
      1. Make a POST request to http://www.nysed.gov/coms/rp090/IRPS2A to get a web page listing
      all QC programs, and extract the numeric program codes and Unit Codes.
      2. Make a GET request to http://www.nysed.gov/COMS/RP090/IRPSL3 for each of the program
      codes retrieved from Step 1, and analyze each page returned to extract the information needed
      to generate the spreadsheet to generate for output.

January 2016:
Deductions, based on examination of the program_codes.out file:
  M/A is for Multiple Awards: the same program code may be associated with more than one degree
  M/I is for Multiple Institutions: the same program may be associated with more than one
      institution, which may use different HEGIS codes and/or awards. If the institution award
      is "NON GRANTING" there is no HEGIS code for that institution.

Input file structure:
  One program code line
  Zero or more M/I or M/A lines
  For each award:
    For award (BA, MA, etc.)
      Program first and last registration dates
      Certificate, etc.
      Program registration dates
      Financial Aid
      Accreditation

Algorithm:
  Read a line, and extract line_type (see the Line Type enum)
  Check previous line_type against this line_type to make sure a valid sequence is taking place.
  Then go to the particular algorithm step given below based on this line_type.

  program:                Extract program_code, program_name, hegis_code, award, institution;
                          check for duplicate program code;
  multiple_awards:        Extract program_name, HEGIS, award, institution. Add to respective sets.
  multiple_instituitions: Extract HEGIS, award, institution.
  for_award:              Extract award.
  certificate_etc:        Extract certificate tuple {name, type, date} if there is one.
  program first:          Extract first and last registration dates
  financial_aid:          Extract three booleans.
  accreditation:          Extract text, if any.

Expect lines in a detail report in the following sequence:
  Always:
    Program Code No. [Program Name] [HEGIS code] [Degree] [Institution]
  Optional:
    M/I WITH - - - > [Degree] [Institution]
    M/A [Program Name] [HEGIS code] [Degree] [Institution]
  Always:
    FOR AWARD -- [Degree]
    CERTIFICATES/LICENSES TITLES AND TYPES: [text]
    PROGRAM FIRST REG DATE: [first_reg_date] LAST REG DATE: [last_reg_date]
    PROGRAM FINANCIAL AID ELIGIBILITY: [TAP: YES|NO] [APTS: YES|NO] [VVTA: YES|NO]
    PROGRAM PROFESSIONAL ACCREDITATION: [text]

"""
import sys
import re
import argparse

from enum import Enum
from collections import namedtuple

import requests
from lxml.html import document_fromstring
import cssselect

from openpyxl import Workbook
from openpyxl.styles import Font

__author__ = 'Christopher Vickery'
__version__ = 'April 2019'

parser = argparse.ArgumentParser()
parser.add_argument('-d', '--debug', action='store_true', default=False)
args = parser.parse_args()


def detail_lines(all_lines, program_code):
  """ Filter out unwanted lines from a details web page for a program code, and generate the others.
  """
  lines = all_lines.splitlines()
  for line in lines:
    if re.search(f'$program_code|FOR AWARD|PROGRAM|PROGRAM|CERTIFICATE|QUEENS',
                 line):
      if args.debug:
        print(line)
      yield(line.replace('<H4><PRE>', '').strip())


def fix_title(str):
  """ Create a better titlecase string, taking specifics of this dataset into account.
  """
  return (str.strip()
             .title()
             .replace('Mhc', 'MHC')
             .replace('\'S', '’s')
             .replace('1St', '1st')
             .replace('6Th', '6th')
             .replace(' And ', ' and ')
             .replace('Cuny', 'CUNY'))


""" Input line types, determined from the first token on the line.
"""
line_type = Enum('Line Type', """
                 program
                 program_dates
                 multiple_awards
                 multiple_institutions
                 for_award
                 certificate_etc
                 financial_aid
                 accreditation""")


Key = namedtuple('Key', """
                 program_code
                 program_name
                 hegis_code
                 award
                 institution""")
Record = namedtuple('Record', """
                    first_date
                    last_date
                    cert_name
                    cert_type
                    cert_date
                    tap_eligible
                    apts_eligible
                    vvta_eligible
                    accreditation
                    unit_code""")
records = dict()

MAW_line = namedtuple('MAW_line', 'name, hegis, award, institution')
Certs_etc = namedtuple('Certs_etc', 'name type date')
Certs_etc.__new__.__defaults__ = (None,) * len(Certs_etc._fields)

Financial_aid = namedtuple('Financial_aid', 'tap apts vvta')
Financial_aid.__new__.__defaults__ = (False,) * len(Financial_aid._fields)

# Fetch the web page with all QC program codes and their unit codes
print('Fetching program code ids and their unit codes', file=sys.stderr)
r = requests.post('http://www.nysed.gov/coms/rp090/IRPS2A', data={'SEARCHES': '1',
                                                                  'instid': '33400'})
html_document = document_fromstring(r.content)
# the program codes and unit codes are inside H4 elements, which alternate: first a program code
# line then, later, a separate unit code. For example:
#   PROGRAM CODE  : 36256 - ...
#   UNIT CODE     : OCUE
h4s = [h4.text_content() for h4 in html_document.cssselect('h4') if 'CODE' in h4.text_content()]
# [Waiting for assignment expressions in python 3.8 (PEP 572)]
# h4s = [h4_line for h4 in html_document.cssselect('h4') if 'CODE' in h4_line := h4.text_content()]

unit_codes = dict()
previous_code_type = object()
previous_code_value = None
for h4 in h4s:
  if args.debug:
    print(h4)
  matches = re.search(r'(PROGRAM|UNIT)\s+CODE\s+:\s+(\S+)', h4)
  this_code_type = matches.group(1)
  this_code_value = matches.group(2)
  if this_code_type is None or this_code_type == previous_code_type:
    sys.exit(f'Invalid h4 sequence: code type is {this_code_type}')
  if this_code_type == 'UNIT':
    unit_codes[previous_code_value] = this_code_value
  previous_code_type = this_code_type
  previous_code_value = this_code_value

print(f'Found {len(unit_codes.keys())} distinct program codes\nFetching details...',
      file=sys.stderr)

prog_code_num = 0
this_type = line_type.accreditation
prev_type = object()
num_dupes = 0
program_codes = set()

for program_code in unit_codes.keys():
  unit_code = unit_codes[program_code]
  r = requests.get(f'http://www.nysed.gov/COMS/RP090/IRPSL3?PROGCD={program_code}')
  prog_code_num += 1
  print(f'Program_code {prog_code_num}/{len(unit_codes.keys())}: {program_code}\r',
        file=sys.stderr,
        end='')
  for line in detail_lines(r.text, program_code):
    # Use the first token on a line to determine the type of line
    prev_type = this_type
    tokens = line.split()
    token = tokens[0]

    if token.isdecimal():
      # There is a number (Program Code No.) at the beginning, so this is a line_type.program
      this_type = line_type.program
      if prev_type is not line_type.accreditation:
        print(f'line {prog_code_num}: unexpected program code line')
      # Extract program_code, program_name, hegis_code, award, institution. Note: allow a single
      # space in the award string ---------------------> vvvvvvvvv
      matches = re.match(r'\s*(\d+)(.+)(\d{4}\.\d{2})\s+(\S+\s?\S*)\s+(.+)', line)
      if matches is None:
        sys.exit(f'Unable to parse program code line for program code {program_code}:\n{line}')
      if program_code != token or program_code != matches.group(1):
        sys.exit(f'Mismatched program codes: "{program_code}" - "{token}" = "{matches.group(1)}"')
      program_name = fix_title(matches.group(2))
      program_hegis = matches.group(3)
      program_award = matches.group(4).strip()
      program_institution = fix_title(matches.group(5))
      # Check for duplicated program code;
      if program_code in program_codes:
        if args.debug:
          print('line {}: duplicate program code: {}'.format(prog_code_num, program_code))
        num_dupes = num_dupes + 1
      program_codes.add(program_code)

      # Initialize structures for this program code
      multiple_award_lines = set()
      multiple_award_awards = {program_award}
      multiple_institution_line = False
      cert = Certs_etc()  # default in case there isn't a certificates line

      if args.debug:
        print(f'{program_code}: "{program_name}" {program_hegis}'
              f' {program_award} "{program_institution}"')

    if token == 'M/A':
      # line_type.multiple_awards
      #
      # There may be multiple M/A (multiple award) lines per program code
      this_type = line_type.multiple_awards
      if prev_type is not line_type.program and \
         prev_type is not line_type.multiple_awards and \
         prev_type is not line_type.multiple_institutions:
        print('line {}: unexpected M/A line'.format(prog_code_num), file=sys.stderr)
      # Extract name, hegis, award, and institution
      matches = re.match(r'\s*M/A(.+)(\d{4}\.\d{2})\s+(\S+\s?\S*)\s+(.+)', line)
      if matches is None:
        sys.exit(f'Unable to parse M/A line for program code {program_code}:\n{line}')
      multiple_award_name = fix_title(matches.group(1))
      multiple_award_hegis = matches.group(2)
      multiple_award_award = matches.group(3).strip()
      multiple_award_institution = fix_title(matches.group(4))
      multiple_award_line = MAW_line(multiple_award_name,
                                     multiple_award_hegis,
                                     multiple_award_award,
                                     multiple_award_institution)
      if multiple_award_line in multiple_award_lines:
        print('line {}: duplicate M/A ({}) for program code {}'
              .format(prog_code_num, multiple_award_line, program_code), file=sys.stderr)
      multiple_award_lines.add(multiple_award_line)
      multiple_award_awards.add(multiple_award_award)  # For quick lookup
      if args.debug:
        print('M/A:',
              multiple_award_name,
              multiple_award_hegis,
              multiple_award_award,
              multiple_award_institution)

    if token == 'M/I':
      # line_type.multiple_institutions
      #
      # Observation: there is never more than one M/I (multiple institutions)
      # line with an award per program code. There might be multiple M/I lines
      # but no more than one of them has an award. If this ever changes, this
      # section of the code will have to be updated to work like M/A lines.
      this_type = line_type.multiple_institutions
      if prev_type is not line_type.program and \
         prev_type is not line_type.multiple_awards and\
         prev_type is not line_type.multiple_institutions:
        print('line {}: unexpected M/I line'.format(prog_code_num), file=sys.stderr)
      # Extract hegis, award, institution
      #   But there is no hegis if the award is NOT-GRANTING
      if 'NOT-GRANTING' in line:
        multiple_institution_hegis = None
        multiple_institution_award = None
        multiple_institution_institution = fix_title(line.split('NOT-GRANTING ')[1])
      else:
        matches = re.match(r'\s*M/A\s+WITH.+>\s+(\d{4}.\d{2})(\S+\s/\S+) (.*)', line)
        if matches is None:
          sys.exit(f'Unable to parse M/I line for program code {program_code}:\n{line}')
        multiple_institution_hegis = matches.group(1)
        multiple_institution_award = matches.group(2)
        if multiple_institution_award:
          if multiple_institution_line:
            print(f'line {prog_code_num}: second M/I line for program code {program_code}:\n{line}',
                  file=sys.stderr)
          multiple_institution_institution = fix_title(matches.group(3))
          multiple_institution_line = True
      if args.debug:
        print('M/I:',
              multiple_institution_hegis,
              multiple_institution_award,
              multiple_institution_institution)

    if token == 'FOR':
      # line_type.for_award
      #
      this_type = line_type.for_award
      if prev_type is not line_type.program and \
         prev_type != line_type.multiple_awards and \
         prev_type != line_type.multiple_institutions and\
         prev_type != line_type.accreditation:
        print('line {}: unexpected for_award line'.format(prog_code_num), file=sys.stderr)
      # Extract award (is it in the set of awards extracted from program and m/a lines?)
      for_award = re.match(r'\s*FOR AWARD\s*--(.*)', line).group(1).strip()
      if for_award != program_award and for_award not in multiple_award_awards:
        # fatal, but why? how does multiple_award_awards know until it sees this?
        print('line {}: FOR award ({}) not in multiple_award_awards ({})'
              .format(prog_code_num, for_award, multiple_award_awards), file=sys.stderr)
        exit()

    if token == 'CERTIFICATE/LICENSES':
      # line_type.certificate_etc
      #
      this_type = line_type.certificate_etc
      if prev_type != line_type.for_award:
        print('line{}: unexpected certificate_etc line'.format(prog_code_num), file=sys.stderr)
      # Extract certificate tuple {name, type, date} if there is one.
      cert_info = line[47:].strip()
      if cert_info.startswith('NONE'):
        cert = Certs_etc()
      else:
        cert = Certs_etc(cert_info[0:18].strip(), cert_info[18:28].strip(), cert_info[28:].strip())

    if token == 'PROGRAM' and tokens[1] == 'FINANCIAL':
      # line_type.financial_aid
      #
      this_type = line_type.financial_aid
      if prev_type != line_type.certificate_etc and\
         prev_type != line_type.for_award:
        print('line {}: unexpected financial_aid line'.format(prog_code_num), file=sys.stderr)
      # Extract three booleans.
      if_tap = line[54:57] == 'YES'
      if_apts = line[66:69] == 'YES'
      if_vvta = line[77:80] == 'YES'
      fin_aid = Financial_aid(if_tap, if_apts, if_vvta)

    if token == 'PROGRAM' and tokens[1] == 'PROFESSIONAL':
      # line_type.accreditation
      #
      this_type = line_type.accreditation
      if prev_type != line_type.financial_aid:
        print('line {}: unexpected accreditation line'.format(prog_code_num), file=sys.stderr)
      # Extract text, if any.
      accreditation = line[45:].strip()

      # Observation: every program's section ends with an accreditation line.
      #
      # Generate record(s) for this program.
      #

      # Assemble the data record
      #   cert_name cert_type cert_date tap_eligible apts_eligible vvta_eligible accreditation
      data_record = Record(cert.name, cert.type, cert.date,
                           fin_aid.tap, fin_aid.apts, fin_aid.vvta,
                           accreditation, unit_code)

      # Generate all distinct keys for this award for this program code
      keys = set()

      # Multiple award sets, if any
      if for_award in multiple_award_awards:
        for multiple_award_line in multiple_award_lines:
          if multiple_award_line.award == for_award:
            key = Key(program_code,
                      multiple_award_line.name,
                      multiple_award_line.hegis,
                      multiple_award_line.award,
                      multiple_award_line.institution)
            if key in records.keys():
              if args.debug:
                print('Duplicate key while procssing M/A for program code {}:\n  {}'
                      .format(program_code, key))
              if records[key] != data_record:
                print('Duplicated maw key with different data: {}\n  {}\n  {}'.format(
                    key, records[key], data_record), file=sys.stderr)
            else:
              records[key] = data_record

      # Multiple institutions, if any
      if multiple_institution_line and multiple_institution_award == for_award:
        key = Key(program_code,
                  program_name,
                  multiple_institution_hegis,
                  multiple_institution_award,
                  multiple_institution_institution)
        if key in records.keys():
          if args.debug:
            print('Duplicate key while processing M/I for program code {}:\n  {}'
                  .format(program_code, key))
          if records[key] != data_record:
            print('Duplicated M/I key with different data: {}\n  {}\n  {}'.format(
                key, records[key], data_record), file=sys.stderr)
        else:
          records[key] = data_record

      # Programs without multiple institutions or awards
      if program_award == for_award:
        key = Key(program_code,
                  program_name,
                  program_hegis,
                  program_award,
                  program_institution)
        if key in records.keys():
          if args.debug:
            print('Duplicate key while processing program code {}:\n  {}'
                  .format(program_code, key))
          if records[key] != data_record:
            print('Duplicated program key with different data: {}\n  {}\n  {}'.format(
                key, records[key], data_record))
        else:
          records[key] = data_record

print(file=sys.stderr)
exit()

# Generate spreadsheet
#
wb = Workbook()
ws = wb.active
ws.title = 'Academic Programs'
row = 1
headers = ('Program Code', 'Program Name', 'HEGIS', 'Award', 'Institution',
           'Certificate Name', 'Certificate Type', 'Certificate Date',
           'TAP Eligible', 'APTS Eligible', 'VVTA Eligible')
bold = Font(bold=True)
for col in range(len(headers)):
  cell = ws.cell(row=row, column=col + 1)
  cell.value = headers[col]
  cell.font = bold
for key in sorted(records.keys()):
  row = row + 1
  for col in range(len(key)):
    cell = ws.cell(row=row, column=col + 1)
    cell.value = key[col]
  for col in range(len(records[key])):
    cell = ws.cell(row=row, column=col + len(key) + 1)
    cell.value = records[key][col]
wb.save('program_codes.xlsx')
print('Processed {} lines.\nFound {} distinct program codes and {} duplicate program codes'
      .format(prog_code_num, len(program_codes), num_dupes))
print('Generated', len(records), 'records')
